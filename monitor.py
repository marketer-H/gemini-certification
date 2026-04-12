#!/usr/bin/env python3
"""
도서 평점 모니터: 교보문고 / 예스24 / 알라딘
평점이 threshold 아래로 내려가면 알림 발송
"""

import json
import os
import re
import sys
import smtplib
import threading
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from datetime import datetime

import requests

warnings.filterwarnings("ignore")

# ─── 파일 경로 ────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
CONFIG_FILE = BASE_DIR / "config.json"
ISBNS_FILE  = BASE_DIR / "isbns.txt"
STATE_FILE  = BASE_DIR / "state.json"
CACHE_FILE  = BASE_DIR / "isbn_cache.json"

# ─── 공통 헤더 ────────────────────────────────────────────────
_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
HTML_HEADERS = {
    "User-Agent": _UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Accept-Encoding": "identity",
}
JSON_HEADERS = {**HTML_HEADERS, "Accept": "application/json"}


# ─── 설정 / 상태 로딩 ─────────────────────────────────────────
def load_config() -> dict:
    with open(CONFIG_FILE) as f:
        return json.load(f)

def load_isbns() -> list:
    valid = []
    for line in ISBNS_FILE.read_text().splitlines():
        isbn = line.strip()
        if not isbn:
            continue
        if len(isbn) != 13 or not isbn.isdigit():
            print(f"[SKIP] 잘못된 ISBN: {isbn}")
            continue
        valid.append(isbn)
    return valid

def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {}

def save_state(state: dict):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def load_cache() -> dict:
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}

def save_cache(cache: dict):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ─── 알라딘 ───────────────────────────────────────────────────
def get_aladin_rating(isbn: str) -> tuple:
    """(rating, title, review_count) 반환"""
    url = f"https://www.aladin.co.kr/shop/wproduct.aspx?ISBN={isbn}"
    try:
        r = requests.get(url, headers=HTML_HEADERS, timeout=15)
        if r.status_code != 200:
            return None, "", 0
        m = re.search(r'"ratingValue":\s*"?([0-9]+(?:\.[0-9]+)?)"?', r.text)
        rating = float(m.group(1)) if m else None
        t = re.search(r'<meta property="og:title" content="([^"]+)"', r.text)
        title = t.group(1).strip() if t else ""
        rc = re.search(r'"reviewCount":\s*"?(\d+)"?', r.text)
        review_count = int(rc.group(1)) if rc else 0
        return rating, title, review_count
    except Exception:
        return None, ""


# ─── 예스24 ───────────────────────────────────────────────────
_yes24_lock = threading.Lock()  # Yes24는 동시 요청 차단 → 순차 처리

def get_yes24_rating(isbn: str, session: requests.Session) -> tuple:
    url = f"https://www.yes24.com/Product/Search?query={isbn}&domain=BOOK"
    with _yes24_lock:
        try:
            r = session.get(url, headers=HTML_HEADERS, timeout=15)
            if r.status_code != 200:
                return None, ""
            m = re.search(r'rating_grade.*?yes_b">([0-9]+(?:\.[0-9]+)?)', r.text, re.DOTALL)
            rating = float(m.group(1)) if m else None
            t = re.search(r'class="gd_name"[^>]*>([^<]+)', r.text)
            title = t.group(1).strip() if t else ""
            return rating, title
        except Exception:
            return None, ""


# ─── 교보문고 ─────────────────────────────────────────────────
def get_kyobo_rating(isbn: str, cache: dict, cache_lock: threading.Lock) -> tuple:
    # product ID 조회 (캐시 우선)
    with cache_lock:
        product_id = cache.get(isbn)

    if not product_id:
        url = f"https://search.kyobobook.co.kr/search?keyword={isbn}&gbCode=TOT&target=total"
        try:
            r = requests.get(url, headers={**HTML_HEADERS, "Referer": "https://www.kyobobook.co.kr/"}, timeout=15)
            m = re.search(rf'data-pid="(S\d+)"[^>]*data-bid="{isbn}"', r.text)
            if not m:
                m = re.search(rf'data-bid="{isbn}"[^>]*data-pid="(S\d+)"', r.text)
            if m:
                product_id = m.group(1)
                with cache_lock:
                    cache[isbn] = product_id
        except Exception:
            return None, ""

    if not product_id:
        return None, ""

    # 평점 조회
    try:
        r = requests.get(
            f"https://product.kyobobook.co.kr/api/review/statistics?saleCmdtid={product_id}",
            headers={**JSON_HEADERS, "Referer": f"https://product.kyobobook.co.kr/detail/{product_id}"},
            timeout=15,
        )
        inner = r.json().get("data") or {}
        avg = inner.get("revwRvgrAvg")
        rating = float(avg) if avg is not None else None
    except Exception:
        return None, ""

    # 제목 조회 (캐시 우선)
    title_key = f"_title_{isbn}"
    with cache_lock:
        title = cache.get(title_key, "")

    if not title:
        try:
            rp = requests.get(
                f"https://product.kyobobook.co.kr/detail/{product_id}",
                headers=HTML_HEADERS, timeout=15
            )
            tm = re.search(r'<meta property="og:title" content="([^"]+)"', rp.text)
            title = tm.group(1).strip() if tm else ""
            with cache_lock:
                cache[title_key] = title
        except Exception:
            pass

    return rating, title


# ─── ISBN 1개 처리 ────────────────────────────────────────────
def process_isbn(isbn: str, stores: list, yes24_session: requests.Session,
                 cache: dict, cache_lock: threading.Lock) -> dict:
    """한 ISBN의 모든 서점 평점을 조회. {store: (rating, title)} 반환"""
    results = {}
    for store in stores:
        if store == "aladin":
            rating, title, review_count = get_aladin_rating(isbn)
            results[store] = (rating, title)
            if review_count:
                with cache_lock:
                    cache[f"_reviews_{isbn}"] = review_count
        elif store == "yes24":
            results[store] = get_yes24_rating(isbn, yes24_session)
        elif store == "kyobo":
            results[store] = get_kyobo_rating(isbn, cache, cache_lock)
    return results


# ─── 알림 ─────────────────────────────────────────────────────
def send_slack(webhook_url: str, messages: list):
    if not webhook_url:
        return
    try:
        requests.post(webhook_url, json={"text": "\n\n".join(messages)}, timeout=10)
        print("[Slack 알림 발송 완료]")
    except Exception as e:
        print(f"[Slack 알림 실패] {e}")


def send_email(config: dict, subject: str, text_body: str,
               html_body: str = None, attachment_path: str = None):
    import base64 as b64
    ec = config.get("email", {})
    if not ec.get("enabled") or not ec.get("username"):
        return
    password = ec.get("password") or os.environ.get("SMTP_PASSWORD", "")
    if not password:
        print("[이메일 알림 실패] 비밀번호 없음.")
        return
    try:
        text_body = text_body.replace("\xa0", " ")
        subject_b64 = "=?utf-8?b?" + b64.b64encode(subject.encode("utf-8")).decode("ascii") + "?="

        if attachment_path or html_body:
            outer = MIMEMultipart("mixed")
            outer["From"] = ec["username"]
            outer["To"] = ec["to"]
            outer["Subject"] = subject_b64

            alt = MIMEMultipart("alternative")
            alt.attach(MIMEText(text_body, "plain", "utf-8"))
            if html_body:
                alt.attach(MIMEText(html_body, "html", "utf-8"))
            outer.attach(alt)

            if attachment_path:
                with open(attachment_path, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                fname = Path(attachment_path).name
                fname_b64 = b64.b64encode(fname.encode("utf-8")).decode("ascii")
                part.add_header("Content-Disposition",
                                f'attachment; filename="=?utf-8?b?{fname_b64}?="')
                outer.attach(part)

            raw = outer.as_bytes()
        else:
            # 첨부/HTML 없을 때 기존 방식 유지
            body_b64 = b64.b64encode(text_body.encode("utf-8")).decode("ascii")
            raw = "\r\n".join([
                f"From: {ec['username']}",
                f"To: {ec['to']}",
                f"Subject: {subject_b64}",
                "MIME-Version: 1.0",
                'Content-Type: text/plain; charset="utf-8"',
                "Content-Transfer-Encoding: base64",
                "",
                body_b64,
            ]).encode()

        with smtplib.SMTP(ec["smtp_server"], ec["smtp_port"]) as server:
            server.starttls()
            server.login(ec["username"], password)
            server.sendmail(ec["username"], [ec["to"]], raw)
        print("[이메일 알림 발송 완료]")
    except Exception as e:
        print(f"[이메일 알림 실패] {e}")


# ─── AI 권고 / HTML / Excel ───────────────────────────────────

def generate_ai_recommendations(below: list, contexts: dict = None, batch_size: int = 15) -> dict:
    """Claude CLI(-p)로 도서별 대응 권고 생성 (배치 처리). {'isbn_store': '권고문'} 반환"""
    import subprocess
    if not below:
        return {}
    all_recs = {}
    contexts = contexts or {}

    for i in range(0, len(below), batch_size):
        batch = below[i:i + batch_size]
        book_lines = []
        for r, store, isbn, title, url in batch:
            ctx = contexts.get(isbn, {})
            all_ratings = ctx.get("all_ratings", {})
            prev_r = ctx.get("prev_rating")
            review_count = ctx.get("review_count", 0)

            trend = ""
            if prev_r is not None and prev_r != r:
                trend = f"이전 {prev_r:.1f}→현재 {r:.1f} ({'하락' if r < prev_r else '상승'})"
            elif prev_r == r:
                trend = f"변동 없음 ({r:.1f})"

            line = f"- {title} (ISBN: {isbn}, 서점: {store}, 현재 평점: {r:.1f})"
            if trend:
                line += f"\n  평점 변동: {trend}"
            if review_count:
                line += f"\n  리뷰 수: {review_count}개"
            book_lines.append(line)

        prompt = (
            "당신은 IT/컴퓨터 도서 출판사의 마케팅 담당자입니다.\n"
            "아래는 현재 평점이 기준치(9.0) 미만인 도서 목록입니다.\n\n"
            + "\n".join(book_lines) +
            "\n\n각 도서에 대해 지금 당장 실행할 수 있는 마케팅 대응 액션을 1~2문장으로 작성하세요.\n"
            "규칙:\n"
            "- 현황 설명(평점이 낮다, 리뷰가 적다 등)은 쓰지 말 것\n"
            "- 도서별로 서로 다른 구체적 액션을 제안할 것 (예: SNS 홍보, 저자 인터뷰, 독자 이벤트, 커뮤니티 공략, 할인 프로모션, 오탈자 수정 공지 등)\n"
            "- 평점 변동, 리뷰 수, 서점 특성을 고려해 도서마다 다른 전략을 제안할 것\n"
            "- 행동 동사로 시작할 것 (예: '진행하세요', '검토하세요', '발송하세요')\n\n"
            "응답은 반드시 아래 JSON 형식으로만 작성하세요 (설명 없이):\n"
            '{"ISBN_서점": "대응 액션", ...}\n'
            '예: {"9791163030034_aladin": "저자와 협력해 독자 Q&A 이벤트를 알라딘 서평단 대상으로 진행하세요."}'
        )
        try:
            result = subprocess.run(
                ["claude", "-p", prompt],
                capture_output=True, text=True, timeout=120
            )
            text = result.stdout
            # JSON 블록 추출 (코드블록 안에 있을 수도 있음)
            m = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
            if not m:
                m = re.search(r'\{.*\}', text, re.DOTALL)
            if m:
                raw_json = m.group(1) if m.lastindex else m.group()
                try:
                    parsed = json.loads(raw_json)
                    # 키 정규화: 공백 제거, 대소문자 유지
                    for k, v in parsed.items():
                        all_recs[k.strip()] = v
                except json.JSONDecodeError:
                    print(f"[AI 권고] 배치 {i//batch_size + 1} JSON 파싱 실패:\n{raw_json[:200]}")
            else:
                print(f"[AI 권고] 배치 {i//batch_size + 1} JSON 없음:\n{text[:200]}")
            print(f"  배치 {i//batch_size + 1}/{(len(below)-1)//batch_size + 1} 완료 ({len(all_recs)}건)")
        except Exception as e:
            print(f"[AI 권고 생성 실패] 배치 {i//batch_size + 1}: {e}")

    return all_recs


def build_html_email(below: list, recs: dict, threshold: float, now: str) -> str:
    def row_class(r):
        return "critical" if r < 8.5 else "warning"

    rows_html = ""
    for r, store, isbn, title, url in below:
        rec = recs.get(f"{isbn}_{store}", "")
        cls = row_class(r)
        rows_html += (
            f'<tr class="{cls}">'
            f'<td class="rating">{r:.1f}</td>'
            f'<td>{store}</td>'
            f'<td><a href="{url}">{title}</a></td>'
            f'<td class="rec">{rec}</td>'
            f'</tr>\n'
        )

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 14px; color: #333; }}
  h2 {{ color: #2c3e50; }}
  table {{ border-collapse: collapse; width: 100%; margin-top: 16px; }}
  th {{ background: #2c3e50; color: white; padding: 10px 12px; text-align: left; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #ddd; vertical-align: top; }}
  tr.critical td {{ background: #fdecea; }}
  tr.warning td {{ background: #fff8e1; }}
  .rating {{ font-weight: bold; font-size: 1.1em; }}
  tr.critical .rating {{ color: #c0392b; }}
  tr.warning .rating {{ color: #e67e22; }}
  a {{ color: #2980b9; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .rec {{ color: #555; font-style: italic; }}
  .legend {{ margin-top: 12px; font-size: 12px; color: #777; }}
</style>
</head>
<body>
<h2>도서 평점 리포트</h2>
<p>{now} &nbsp;|&nbsp; 평점 <strong>{threshold}</strong> 미만 도서 <strong>{len(below)}건</strong></p>
<table>
  <tr><th>평점</th><th>서점</th><th>도서명</th><th>AI 대응 권고</th></tr>
  {rows_html}
</table>
<p class="legend">
  <span style="background:#fdecea;padding:2px 8px;">■</span> 8.5 미만 (긴급) &nbsp;
  <span style="background:#fff8e1;padding:2px 8px;">■</span> 8.5~{threshold} 미만 (주의)
</p>
</body>
</html>"""


def create_excel(below: list, recs: dict, threshold: float, now: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[Excel 생성 실패] openpyxl 미설치. pip install openpyxl")
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "평점 리포트"

    # 타이틀
    ws.merge_cells("A1:E1")
    ws["A1"] = f"도서 평점 리포트 — {now} | 임계값 {threshold} 미만 {len(below)}건"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 헤더
    headers = ["평점", "서점", "도서명", "링크", "AI 대응 권고"]
    header_fill = PatternFill("solid", fgColor="2C3E50")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    critical_fill = PatternFill("solid", fgColor="FDECEA")
    warning_fill  = PatternFill("solid", fgColor="FFF8E1")
    thin = Border(bottom=Side(style="thin", color="DDDDDD"))

    for row_i, (r, store, isbn, title, url) in enumerate(below, 3):
        fill = critical_fill if r < 8.5 else warning_fill
        rec  = recs.get(f"{isbn}_{store}", "")

        cells = [r, store, title, url, rec]
        for col, val in enumerate(cells, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_i, column=1).font = Font(
            bold=True, color="C0392B" if r < 8.5 else "E67E22"
        )
        ws.cell(row=row_i, column=1).alignment = Alignment(horizontal="center", vertical="top")

    # 열 너비
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 45

    path = str(BASE_DIR / f"평점리포트_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb.save(path)
    return path


def save_dashboard(below: list, recs: dict, threshold: float, now: str, cache: dict = None):
    """대시보드 HTML 파일 저장"""
    cache = cache or {}

    def series_group(title):
        if title.startswith("된다!"):
            return "된다! 시리즈"
        elif title.startswith("Do it!"):
            return "Do it! 시리즈"
        else:
            return "기타"

    def make_rows(items):
        rows = ""
        for r, store, isbn, title, url in items:
            rec = recs.get(f"{isbn}_{store}", "—")
            review_count = cache.get(f"_reviews_{isbn}", "")
            cls = "critical" if r < 8.5 else "warning"
            rows += (
                f'<tr class="{cls}" data-store="{store}">'
                f'<td class="rating">{r:.1f}</td>'
                f'<td><span class="badge {store}">{store}</span></td>'
                f'<td><a href="{url}" target="_blank">{title}</a></td>'
                f'<td class="review-count">{review_count if review_count else "—"}</td>'
                f'<td class="rec">{rec}</td>'
                f'</tr>\n'
            )
        return rows

    def _build_sections(grps, row_fn):
        html_parts = []
        for group, items in grps.items():
            if items:
                table = (
                    '<table class="mainTable">'
                    '<thead><tr><th>평점</th><th>서점</th><th>도서명</th>'
                    '<th style="text-align:center">리뷰 수</th><th>AI 대응 권고</th></tr></thead>'
                    f'<tbody>{row_fn(items)}</tbody></table>'
                )
            else:
                table = "<p style='color:#aaa;font-size:13px;padding:8px 0'>해당 없음</p>"
            html_parts.append(
                f'<div class="section"><h2>{group} <span>{len(items)}건</span></h2>{table}</div>'
            )
        return "\n".join(html_parts)

    critical_count = sum(1 for r, *_ in below if r < 8.5)
    warning_count  = len(below) - critical_count

    # 시리즈별 분류
    groups = {"된다! 시리즈": [], "Do it! 시리즈": [], "기타": []}
    for item in below:
        groups[series_group(item[3])].append(item)

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>도서 평점 대시보드</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         background: #f0f2f5; color: #333; padding: 24px; }}
  h1 {{ font-size: 22px; font-weight: 700; color: #1a1a2e; margin-bottom: 4px; }}
  .meta {{ font-size: 13px; color: #888; margin-bottom: 24px; }}

  .cards {{ display: flex; gap: 16px; margin-bottom: 28px; flex-wrap: wrap; }}
  .card {{ background: white; border-radius: 12px; padding: 20px 28px;
           box-shadow: 0 1px 4px rgba(0,0,0,.08); min-width: 140px; }}
  .card .num {{ font-size: 32px; font-weight: 700; }}
  .card .label {{ font-size: 12px; color: #888; margin-top: 2px; }}
  .card.red .num {{ color: #e74c3c; }}
  .card.orange .num {{ color: #e67e22; }}
  .card.blue .num {{ color: #3498db; }}

  .section {{ background: white; border-radius: 12px; padding: 20px;
              box-shadow: 0 1px 4px rgba(0,0,0,.08); margin-bottom: 20px; }}
  .section h2 {{ font-size: 15px; font-weight: 600; margin-bottom: 14px;
                 padding-bottom: 10px; border-bottom: 1px solid #eee; }}
  .section h2 span {{ font-size: 12px; font-weight: 400; color: #999; margin-left: 8px; }}

  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ text-align: left; padding: 8px 12px; background: #f8f9fa;
        color: #555; font-weight: 600; border-bottom: 2px solid #eee; }}
  td {{ padding: 10px 12px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }}
  tr:last-child td {{ border-bottom: none; }}
  tr.critical td {{ background: #fff5f5; }}
  tr.warning td {{ background: #fffbf0; }}
  tr:hover td {{ filter: brightness(0.97); }}

  .rating {{ font-weight: 700; font-size: 15px; width: 50px; }}
  tr.critical .rating {{ color: #e74c3c; }}
  tr.warning .rating {{ color: #e67e22; }}

  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 20px;
            font-size: 11px; font-weight: 600; }}
  .badge.aladin {{ background: #e8f4fd; color: #2980b9; }}
  .badge.yes24  {{ background: #fef9e7; color: #d68910; }}
  .badge.kyobo  {{ background: #eafaf1; color: #27ae60; }}

  a {{ color: #2980b9; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .rec {{ color: #555; line-height: 1.5; max-width: 380px; }}
  .review-count {{ text-align: center; color: #888; font-size: 12px; width: 60px; }}

  input[type=text] {{ width: 100%; padding: 8px 12px; border: 1px solid #ddd;
                      border-radius: 8px; font-size: 13px; margin-bottom: 12px; }}

  .filters {{ display: flex; gap: 8px; margin-bottom: 14px; flex-wrap: wrap; }}
  .filter-btn {{ padding: 5px 14px; border-radius: 20px; border: 1px solid #ddd;
                 background: white; font-size: 12px; font-weight: 600; cursor: pointer;
                 transition: all .15s; }}
  .filter-btn:hover {{ filter: brightness(0.95); }}
  .filter-btn.active {{ color: white; border-color: transparent; }}
  .filter-btn[data-store="all"].active {{ background: #555; }}
  .filter-btn[data-store="aladin"].active {{ background: #2980b9; }}
  .filter-btn[data-store="yes24"].active  {{ background: #d68910; }}
  .filter-btn[data-store="kyobo"].active  {{ background: #27ae60; }}
</style>
</head>
<body>
<h1>도서 평점 대시보드</h1>
<p class="meta">마지막 업데이트: {now} &nbsp;|&nbsp; 임계값: {threshold} 미만</p>

<div class="cards">
  <div class="card blue"><div class="num">{len(below)}</div><div class="label">전체 미만 도서</div></div>
  <div class="card red"><div class="num">{critical_count}</div><div class="label">긴급 (8.5 미만)</div></div>
  <div class="card orange"><div class="num">{warning_count}</div><div class="label">주의 (8.5~{threshold})</div></div>
</div>

<input type="text" id="search" placeholder="도서명 검색..." onkeyup="applyFilter()">

<div class="filters">
  <button class="filter-btn active" data-store="all" onclick="setStore('all')">전체</button>
  <button class="filter-btn" data-store="aladin" onclick="setStore('aladin')">알라딘</button>
  <button class="filter-btn" data-store="yes24"  onclick="setStore('yes24')">예스24</button>
  <button class="filter-btn" data-store="kyobo"  onclick="setStore('kyobo')">교보문고</button>
</div>

<p style="font-size:12px;color:#aaa;margin-bottom:16px;">총 <span id="count">{len(below)}</span>건 표시 중</p>

{_build_sections(groups, make_rows)}

<script>
let activeStore = 'all';

function setStore(store) {{
  activeStore = store;
  document.querySelectorAll('.filter-btn').forEach(btn => {{
    btn.classList.toggle('active', btn.dataset.store === store);
  }});
  applyFilter();
}}

function applyFilter() {{
  const q = document.getElementById('search').value.toLowerCase();
  let visible = 0;
  document.querySelectorAll('.mainTable tbody tr').forEach(tr => {{
    const storeMatch = activeStore === 'all' || tr.dataset.store === activeStore;
    const textMatch  = tr.textContent.toLowerCase().includes(q);
    const show = storeMatch && textMatch;
    tr.style.display = show ? '' : 'none';
    if (show) visible++;
  }});
  document.getElementById('count').textContent = visible;
}}
</script>
</body>
</html>"""

    path = BASE_DIR / "dashboard.html"
    path.write_text(html, encoding="utf-8")
    print(f"[대시보드 저장] {path}")
    return str(path)


# ─── 메인 ─────────────────────────────────────────────────────
def clean_title(title: str, isbn: str) -> str:
    """도서명에서 불필요한 접미사 제거"""
    if not title or title == isbn:
        return isbn
    # " | ..." 이후 제거 (알라딘: "제목 | 시리즈 | 저자")
    title = title.split(" | ")[0].strip()
    # " - 교보문고" 등 제거
    for suffix in [" - 교보문고", " - 예스24", " - 알라딘"]:
        if title.endswith(suffix):
            title = title[:-len(suffix)].strip()
    return title


def report():
    """state.json 기반으로 현재 9.5 아래인 도서 목록 출력"""
    config    = load_config()
    threshold = config["threshold"]
    state     = load_state()
    cache     = load_cache()
    stores    = config.get("stores", ["aladin", "yes24", "kyobo"])

    rows = []
    for isbn, isbn_state in state.items():
        for store in stores:
            rating = isbn_state.get(store)
            if rating is not None and rating < threshold:
                title_key = f"_title_{isbn}"
                raw_title = cache.get(title_key) or isbn
                title = clean_title(raw_title, isbn)
                rows.append((rating, store, isbn, title))

    if not rows:
        print(f"\n임계값({threshold}) 아래인 도서 없음.\n")
        return

    rows.sort(key=lambda x: x[0])  # 낮은 평점 순

    # 콘솔 출력
    print(f"\n{'='*60}")
    print(f"현재 평점 {threshold} 미만 도서 ({len(rows)}건)")
    print(f"{'='*60}")
    print(f"{'평점':>5}  {'서점':<8}  {'ISBN':<15}  도서명")
    print(f"{'-'*60}")
    for rating, store, isbn, title in rows:
        print(f"{rating:>5.1f}  {store:<8}  {isbn:<15}  {title}")
    print(f"{'='*60}\n")

    # 이메일 발송
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    below = []
    for rating, store, isbn, title in rows:
        store_url = {
            "aladin": f"https://www.aladin.co.kr/shop/wproduct.aspx?ISBN={isbn}",
            "yes24":  f"https://www.yes24.com/Product/Search?query={isbn}&domain=BOOK",
            "kyobo":  f"https://product.kyobobook.co.kr/detail/{cache.get(isbn, isbn)}",
        }.get(store, "")
        below.append((rating, store, isbn, title, store_url))

    print("\nAI 대응 권고 생성 중...")
    contexts = {}
    for r, store, isbn, title, url in below:
        if isbn not in contexts:
            all_ratings = {s: state.get(isbn, {}).get(s) for s in stores if state.get(isbn, {}).get(s) is not None}
            review_count = cache.get(f"_reviews_{isbn}", 0)
            contexts[isbn] = {"all_ratings": all_ratings, "prev_rating": None, "review_count": review_count}
    recs = generate_ai_recommendations(below, contexts=contexts)

    lines = [f"[{now_str}] 현재 평점 {threshold} 미만 도서 ({len(below)}건)\n"]
    for r, store, isbn, title, url in below:
        rec = recs.get(f"{isbn}_{store}", "")
        lines.append(f"{r:.1f}  {store:<8}  {title}\n{url}" + (f"\n권고: {rec}" if rec else ""))
    text_body = "\n\n".join(lines)

    html_body  = build_html_email(below, recs, threshold, now_str)
    excel_path = create_excel(below, recs, threshold, now_str)
    subject    = f"[도서 평점 리포트] {now_str} | 평점 {threshold} 미만 {len(below)}건"
    send_email(config.get("notification", {}), subject, text_body,
               html_body=html_body, attachment_path=excel_path if excel_path else None)
    if excel_path and Path(excel_path).exists():
        Path(excel_path).unlink()


def run():
    init_mode = "--init" in sys.argv

    config    = load_config()
    threshold = config["threshold"]
    isbns     = load_isbns()
    state     = load_state()
    cache     = load_cache()
    stores    = config.get("stores", ["aladin", "yes24", "kyobo"])
    workers   = config.get("workers", 8)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    mode_label = " [초기화 모드]" if init_mode else ""
    print(f"\n{'='*60}")
    print(f"도서 평점 모니터 실행: {now}{mode_label}")
    print(f"대상 도서: {len(isbns)}권 | 임계값: {threshold} | 병렬: {workers}개")
    print(f"{'='*60}\n")

    # 이전 상태 저장 (AI 권고에 변동 정보 활용)
    prev_state = {isbn: dict(s) for isbn, s in state.items()}

    # Yes24 세션 1회 초기화 (쿠키 획득)
    yes24_session = requests.Session()
    try:
        yes24_session.get("https://www.yes24.com/", headers=HTML_HEADERS, timeout=15)
    except Exception:
        pass

    cache_lock  = threading.Lock()
    print_lock  = threading.Lock()
    state_lock  = threading.Lock()
    alerts      = []
    alerts_lock = threading.Lock()
    done        = [0]

    def handle_isbn(isbn: str):
        results = process_isbn(isbn, stores, yes24_session, cache, cache_lock)

        # 알라딘에서 가져온 제목을 다른 서점 알림에도 공유
        shared_title = ""
        for store in stores:
            _, t = results.get(store, (None, ""))
            if t and t != isbn:
                shared_title = t
                break

        # 제목 캐시에 없으면 저장
        title_key = f"_title_{isbn}"
        with cache_lock:
            if not cache.get(title_key) and shared_title:
                cache[title_key] = shared_title

        isbn_alerts = []
        with state_lock:
            isbn_state = state.setdefault(isbn, {})

        for store in stores:
            rating, title = results.get(store, (None, ""))
            book_name = clean_title(title or shared_title, isbn)

            with state_lock:
                prev = isbn_state.get(store)
                isbn_state[store] = rating if rating is not None else prev

            if rating is None:
                continue

            if not init_mode and rating < threshold and (prev is None or rating < prev):
                store_url = {
                    "aladin": f"https://www.aladin.co.kr/shop/wproduct.aspx?ISBN={isbn}",
                    "yes24":  f"https://www.yes24.com/Product/Search?query={isbn}&domain=BOOK",
                    "kyobo":  f"https://product.kyobobook.co.kr/detail/{cache.get(isbn, isbn)}",
                }[store]
                isbn_alerts.append(
                    f"⚠️ 평점 하락 감지!\n"
                    f"  서점: {store}\n"
                    f"  도서: {book_name} (ISBN: {isbn})\n"
                    f"  현재 평점: {rating:.1f} (이전: {f'{prev:.1f}' if prev is not None else '최초 확인'}) / 임계값: {threshold}\n"
                    f"  URL: {store_url}"
                )

        with state_lock:
            state[isbn] = isbn_state
            done[0] += 1
            n = done[0]

        rating_str = " | ".join(
            f"{s}: {results[s][0]:.1f}" if results.get(s) and results[s][0] is not None else f"{s}: -"
            for s in stores
        )
        alert_mark = " *** 알림 ***" if isbn_alerts else ""
        with print_lock:
            print(f"[{n}/{len(isbns)}] {isbn}  {rating_str}{alert_mark}")

        if isbn_alerts:
            with alerts_lock:
                alerts.extend(isbn_alerts)

    # 병렬 실행
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(handle_isbn, isbn): isbn for isbn in isbns}
        for f in as_completed(futures):
            try:
                f.result()
            except Exception as e:
                with print_lock:
                    print(f"  [오류] {futures[f]}: {e}")

    save_state(state)
    save_cache(cache)

    print(f"\n{'='*60}")
    if init_mode:
        print("초기화 완료. 이제 'python3 monitor.py' 로 모니터링을 시작하세요.")
    else:
        # 전체 임계값 미만 도서 목록 수집 (현재 isbns.txt에 있는 것만)
        isbn_set = set(isbns)
        below = []
        for isbn, isbn_state in state.items():
            if isbn not in isbn_set:
                continue
            for store in stores:
                r = isbn_state.get(store)
                if r is not None and r < threshold:
                    title_key = f"_title_{isbn}"
                    raw_title = cache.get(title_key) or isbn
                    title = clean_title(raw_title, isbn)
                    store_url = {
                        "aladin": f"https://www.aladin.co.kr/shop/wproduct.aspx?ISBN={isbn}",
                        "yes24":  f"https://www.yes24.com/Product/Search?query={isbn}&domain=BOOK",
                        "kyobo":  f"https://product.kyobobook.co.kr/detail/{cache.get(isbn, isbn)}",
                    }.get(store, "")
                    below.append((r, store, isbn, title, store_url))
        below.sort(key=lambda x: x[0])

        if below:
            print(f"현재 평점 {threshold} 미만 도서 ({len(below)}건):\n")
            for r, store, isbn, title, url in below:
                print(f"  {r:.1f}  {store:<8}  {isbn}  {title}")
        else:
            print(f"현재 평점 {threshold} 미만 도서 없음.")

        # 이메일 발송
        notif = config.get("notification", {})
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        if below:
            # AI 권고용 context 빌드 (서점 간 비교, 이전 대비 변동, 리뷰 수)
            contexts = {}
            for r, store, isbn, title, url in below:
                if isbn not in contexts:
                    all_ratings = {s: state[isbn].get(s) for s in stores if state[isbn].get(s) is not None}
                    prev_r = prev_state.get(isbn, {}).get(store)
                    review_count = cache.get(f"_reviews_{isbn}", 0)
                    contexts[isbn] = {
                        "all_ratings": all_ratings,
                        "prev_rating": prev_r,
                        "review_count": review_count,
                    }

            print("\nAI 대응 권고 생성 중...")
            recs = generate_ai_recommendations(below, contexts=contexts)

            # 텍스트 본문
            lines = [f"[{now_str}] 현재 평점 {threshold} 미만 도서 ({len(below)}건)\n"]
            for r, store, isbn, title, url in below:
                rec = recs.get(f"{isbn}_{store}", "")
                lines.append(f"{r:.1f}  {store:<8}  {title}\n{url}" + (f"\n권고: {rec}" if rec else ""))
            text_body = "\n\n".join(lines)

            # HTML + Excel
            html_body = build_html_email(below, recs, threshold, now_str)
            excel_path = create_excel(below, recs, threshold, now_str)

            subject = f"[도서 평점 리포트] {now_str} | 평점 {threshold} 미만 {len(below)}건"
            send_slack(notif.get("slack_webhook", ""), [text_body])
            send_email(notif, subject, text_body, html_body=html_body,
                       attachment_path=excel_path if excel_path else None)

            # 임시 Excel 파일 삭제
            if excel_path and Path(excel_path).exists():
                Path(excel_path).unlink()

            # 대시보드 HTML 저장
            save_dashboard(below, recs, threshold, now_str, cache=cache)
        else:
            print("이메일 발송 없음 (미만 도서 없음).")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    if "--report" in sys.argv:
        report()
    else:
        run()
